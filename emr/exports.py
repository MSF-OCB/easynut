# -*- coding: utf-8 -*-
import json
import os
from collections import OrderedDict

from django.conf import settings
from django.utils.encoding import force_text

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, NamedStyle, PatternFill
from openpyxl.utils import column_index_from_string, coordinate_from_string, get_column_letter

from .models import MSF_ID_VERBOSE_NAME, DATE_VERBOSE_NAME, RE_DATA_SLUG_VALIDATION, DynamicRegistry
from .utils import DataDb, insert_filename_pre_extension, now_for_filename, xlsx_download_response_factory


# Value allowing to skip a cell while continuing to read config of other columns.
DATA_SLUG_EMPTY_CELL = "#"


# BASE CLASSES ================================================================

class AbstractExportExcel(object):

    DEFAULT_FILENAME = "easynut-export.xlsx"  # Default name of the file.

    # Minimum number of cols to apply formatting to.
    APPLY_STYLE_MIN_NUM_COLS = 30

    # Available named styles.
    STYLE_HEADING = "heading"  # For heading rows.
    STYLE_VERBOSE = "advanced"  # For advanced information (cf. ``VERBOSE``).
    STYLE_MSF_ID = "msf_id"  # For the special field ``MSF ID``.
    STYLE_DATE = "date"  # For the special field ``Date``.

    # Named styles config.
    HEADING_BG_COLOR = "DDDDFF"
    VERBOSE_FG_COLOR = "999999"
    MSF_ID_BG_COLOR = "FFDDDD"
    DATE_BG_COLOR = "DDFFDD"

    def __init__(self):
        self.book = None  # The Excel workbook.
        self.filename = None  # The name of the file.

        self.styles = {}  # Named styles for formatting Excel cells.

        # Initialize the file name.
        self.init_filename()

    def apply_style_to_cols(self, style, sheet, min_col, max_col, min_row=None, max_row=None):
        for row_cells in sheet.iter_rows(min_col=min_col, max_col=max_col, min_row=min_row, max_row=max_row):
            for cell in row_cells:
                cell.style = style

    def apply_style_to_rows(self, style, sheet, min_row, max_row, min_col=None, max_col=None):
        for col_cells in sheet.iter_cols(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in col_cells:
                cell.style = style

    def cell_name_to_col_row(self, cell):
        """Convert a cell name into ``(col, row)`` coordinate. E.g. ``B5`` => ``(2, 5)``."""
        col, row = coordinate_from_string(cell)
        col = column_index_from_string(col)
        return col, row

    def cell_col_row_to_name(self, col, row):
        """Convert ``(col, row)`` coordinate into a cell name. E.g. ``(2, 5)`` => ``B5``."""
        col_letter = get_column_letter(col)
        return "{}{}".format(col_letter, row)

    def create_sheet(self, title):
        """Create a new sheet in the current workbook."""
        return self.book.create_sheet(title=title)

    def get_sheet(self, index):
        """Return a sheet identified by its index."""
        return self.book.worksheets[index]

    def get_sheet_names(self):
        """Return the list of sheet names."""
        return self.book.sheetnames

    def get_style(self, name):
        """Return a named style to format an Excel cell."""
        if name not in self.styles:
            self.styles[name] = NamedStyle(name=name)
            if name == self.STYLE_HEADING:
                self.styles[name].font = Font(bold=True)
                self.styles[name].fill = PatternFill(
                    fill_type="solid", start_color=self.HEADING_BG_COLOR, end_color=self.HEADING_BG_COLOR
                )
            elif name == self.STYLE_VERBOSE:
                self.styles[name].font = Font(color=self.VERBOSE_FG_COLOR)
            elif name == self.STYLE_MSF_ID:
                self.styles[name].fill = PatternFill(
                    fill_type="solid", start_color=self.MSF_ID_BG_COLOR, end_color=self.MSF_ID_BG_COLOR
                )
            elif name == self.STYLE_DATE:
                self.styles[name].fill = PatternFill(
                    fill_type="solid", start_color=self.DATE_BG_COLOR, end_color=self.DATE_BG_COLOR
                )
            else:
                return None

        return self.styles[name]

    def init_filename(self):
        """Initialize file name to its default value with "now"."""
        self.filename = insert_filename_pre_extension(self.DEFAULT_FILENAME, now_for_filename())

    def save(self, filename=None):
        """Save the Excel file (under ``settings.EXPORTS_ROOT``)."""
        self._save_common()
        if not filename:
            filename = self.filename

        # Save the file on the file system.
        file_path = os.path.join(settings.EXPORTS_ROOT, filename)
        self.book.save(file_path)

        # Return the path where the file has been saved.
        return file_path

    def save_to_response(self):
        """Save the Excel file in an HTTP response."""
        self._save_common()

        # Create a download response for an Excel (xlsx) file.
        response = xlsx_download_response_factory(self.filename)
        self.book.save(response)
        return response

    def _save_common(self):
        """Common steps for "save" methods."""
        if self.book is None:
            raise Exception("No Excel Workbook to save.")

        # Set active sheet to first one.
        self.book.active = 0


class AbstractExportExcelTemplate(AbstractExportExcel):

    DEFAULT_TEMPLATE = None  # Default Excel template to use.

    def __init__(self, template=None, filename=None):
        super(AbstractExportExcelTemplate, self).__init__()
        self.template = template
        self.filename = filename

        # Config of the loaded template.
        self._config = OrderedDict()

        # Load the template (if not specified, load default one).
        self.load_template(template)

    def load_template(self, template=None):
        """Create the workbook from a template file (located under ``settings.EXPORTS_TEMPLATES_DIR``)."""
        # Set the current template.
        if template is None:
            if self.DEFAULT_TEMPLATE is None:
                raise Exception("No template defined.")
            template = self.DEFAULT_TEMPLATE
        self.template = template

        # Load the template.
        file_path = os.path.join(settings.EXPORTS_TEMPLATES_DIR, self.template)
        self.book = load_workbook(file_path)

        # Read the template config.
        self._init_config()

    def populate(self):
        """Populate the template with data."""
        if self.book is None:
            raise Exception("No template loaded.")
        # Must be implemented in concrete classes. Don't raise ``NotImplemented`` so that they can override this.

    def _init_config(self):
        """
        Read the template config.

        - The config sheet must be the first sheet in the workbook.
        - It is automatically removed.
        - See the concrete class for more specs.
        """
        raise NotImplemented()

    def _sheets_iterator(self, for_config=False):
        """Iterate over configured sheets."""
        for index, config in self._config.iteritems():
            yield index, self.get_sheet(index), config

    def _update_models_fields(self, sheet_index, data_slug):
        """Register the models and fields for a given sheet."""
        # Split the data slug into ``(model_id, field_id)``.
        model_id, field_id = DynamicRegistry.split_data_slug(data_slug)

        # Register the model and field.
        config = self._config[sheet_index]["models_fields"]
        if model_id not in config:
            config[model_id] = []
        if field_id not in config[model_id]:
            config[model_id].append(field_id)


# EXPORTS =====================================================================

class ExportDataModel(AbstractExportExcel):
    """Create an Excel file containing a list of all models and fields with their data slug."""

    DEFAULT_FILENAME = "easynut-data-model.xlsx"  # Default name for the file.
    VERBOSE = True  # Whether to include advanced information.

    # Headings. Use ``None`` to skip a column.
    DATA_SLUGS_HEADINGS = ["Model Name", "Field Name", "Data Slug"]
    DATA_SLUGS_HEADINGS_VERBOSE = [None, "Model ID", "Field ID", "Position", "Type", "Values List"]
    MODELS_HEADINGS = ["Model Name"]
    MODELS_HEADINGS_VERBOSE = ["Model ID"]

    def generate(self):
        """Generate the export."""
        self.book = Workbook()
        DynamicRegistry.load_models_config()

        self._generate_data_slugs()
        self._generate_models()

        return self.book

    def _generate_data_slugs(self):
        """Generate the data slugs sheet."""
        sheet = self.get_sheet(0)
        sheet.title = "Data Slugs"

        headings = self.DATA_SLUGS_HEADINGS
        if self.VERBOSE:
            headings += self.DATA_SLUGS_HEADINGS_VERBOSE
        self._write_headings(sheet, headings)

        self._write_values_data_slugs(sheet)

    def _generate_models(self):
        """Generate the models sheet."""
        sheet = self.create_sheet("Models")

        headings = self.MODELS_HEADINGS
        if self.VERBOSE:
            headings += self.MODELS_HEADINGS_VERBOSE
        self._write_headings(sheet, headings)

        self._write_values_models(sheet)

    def _save_common(self):
        """Common steps for "save" methods."""
        if self.book is None:
            self.generate()
        super(ExportDataModel, self)._save_common()

    def _write_headings(self, sheet, headings):
        """Write heading row."""
        col = 0; row = 1  # NOQA
        for value in headings:
            col += 1
            if value is not None:  # None = skip column.
                sheet.cell(column=col, row=row).value = value

        # Apply heading style and freeze heading row.
        style = self.get_style(self.STYLE_HEADING)
        max_col = max(len(headings), self.APPLY_STYLE_MIN_NUM_COLS)  # Apply to at least a min number of columns.
        self.apply_style_to_rows(style, sheet, min_row=row, max_row=row, max_col=max_col)
        sheet.freeze_panes = sheet.cell(column=1, row=row + 1)

    def _write_values(self, sheet, row, values):
        """Write values on the given row."""
        col = 0
        for value in values:
            col += 1
            if value is not None:  # None = skip column.
                sheet.cell(column=col, row=row).value = value

    def _write_values_data_slugs(self, sheet):
        """Write data slugs values."""
        # Loop over all models and fields config.
        min_row = 2  # Skip heading row.
        row = min_row - 1    # To start the loops with the increment, easier to read.
        for model_id, model_config in DynamicRegistry.models_config.iteritems():
            for field_id, field_config in model_config.fields_config.iteritems():
                row += 1

                # Build values.
                values = [
                    model_config.name,
                    field_config.name,
                    field_config.data_slug,
                ]
                verbose_min_col = len(values) + 1
                if self.VERBOSE:
                    values += [
                        None,
                        model_id,
                        field_id,
                        field_config.position,
                        field_config.kind,
                        json.dumps(field_config.values_list) if field_config.values_list else "",
                    ]
                    verbose_max_col = len(values) + 1

                # Write values.
                self._write_values(sheet, row, values)

                # Apply special fields style.
                style = None
                if field_config.name == MSF_ID_VERBOSE_NAME:
                    style = self.get_style(self.STYLE_MSF_ID)
                elif field_config.name == DATE_VERBOSE_NAME:
                    style = self.get_style(self.STYLE_DATE)
                if style is not None:
                    self.apply_style_to_rows(
                        style, sheet, min_row=row, max_row=row, min_col=1, max_col=verbose_min_col - 1
                    )

                # Apply verbose style.
                style = self.get_style(self.STYLE_VERBOSE)
                self.apply_style_to_cols(
                    style, sheet, min_col=verbose_min_col, max_col=verbose_max_col, min_row=min_row
                )

    def _write_values_models(self, sheet):
        """Write models values."""
        # Loop over all models config.
        min_row = 2  # Skip heading row.
        row = min_row - 1    # To start the loops with the increment, easier to read.
        for model_id, model_config in DynamicRegistry.models_config.iteritems():
            row += 1

            # Build values.
            values = [model_config.name]
            verbose_min_col = len(values) + 1
            if self.VERBOSE:
                values += [model_id]
                verbose_max_col = len(values) + 1

            # Write values.
            self._write_values(sheet, row, values)

            # Apply verbose style.
            style = self.get_style(self.STYLE_VERBOSE)
            self.apply_style_to_cols(style, sheet, min_col=verbose_min_col, max_col=verbose_max_col, min_row=min_row)


class ExportExcelFull(AbstractExportExcel):
    """Create an Excel file containing an export of the whole database."""

    DEFAULT_FILENAME = "easynut-full-export.xlsx"  # Default name for the file.

    def generate(self):
        """Generate the export."""
        self.book = Workbook()
        DynamicRegistry.load_models_config()

        # Loop over all models and fields config.
        for model_id, model_config in DynamicRegistry.models_config.iteritems():
            sheet = self.create_sheet(model_config.name)
            field_ids = []

            # Write headings.
            col = 0; row = 1  # NOQA
            for field_id, field_config in model_config.fields_config.iteritems():
                col += 1
                sheet.cell(column=col, row=row).value = field_config.name
                field_ids.append(field_id)

            # Apply heading style and freeze heading row.
            style = self.get_style(self.STYLE_HEADING)
            num_cols = len(model_config.fields_config)
            self.apply_style_to_rows(style, sheet, min_row=row, max_row=row, num_cols=num_cols)
            sheet.freeze_panes = sheet.cell(column=1, row=row + 1)

            # Write values.
            row = 1
            for model in model_config.objects.all():
                col = 0; row += 1  # NOQA
                for field_id in field_ids:
                    col += 1
                    sheet.cell(column=col, row=row).value = model.get_field_value(field_id)

        # Remove the first sheet.
        self.book.remove_sheet(self.book.active)

        return self.book

    def _save_common(self):
        """Common steps for "save" methods."""
        if self.book is None:
            self.generate()
        super(ExportExcelFull, self)._save_common()


class ExportExcelList(AbstractExportExcelTemplate):
    """Excel export for templates containing a list of records."""

    DEFAULT_FILENAME = "easynut-list.xlsx"  # Default name for the file.
    DEFAULT_TEMPLATE = "easynut-list.xlsx"  # Default template file to use.

    def populate(self):
        """Populate the template with data."""
        super(ExportExcelList, self).populate()

        # Loop over sheets to populate.
        for sheet_index, sheet, config in self._sheets_iterator():
            # Get models and fields needed for this sheet.
            models_fields = self._config[sheet_index]["models_fields"]

            # No config for this sheet? => skip it.
            if len(models_fields) == 0:
                continue

            # Build the SQL statement to get data for this sheet.
            # @TODO: Data are not converted using raw DB queries.
            sql = DynamicRegistry.build_sql(models_fields)

            # Execute the query.
            with DataDb.execute(sql) as c:
                # Loop over data records.
                row = config["start_row"] - 1  # To start the loops with the increment, easier to read.
                for data_row in c.fetchall():
                    # Loop over columns to populate.
                    row += 1
                    for col, data_slug in self._config[sheet_index]["columns"].iteritems():
                        sheet.cell(column=col, row=row).value = data_row[data_slug]

    def _init_config(self):
        """
        Read the template config.

        - See ``AbstractExportExcelTemplate`` for more specs.
        """
        self._config = OrderedDict()  # Reset the config.
        self._init_config_sheets()  # Read config for the sheets to populate.
        self._init_config_columns()  # Read config for the data to populate in each sheet.

    def _init_config_columns(self):
        """
        Read config for the data to populate in each sheet.

        - Data are populated row by row starting at the "starting cell".
        - The first row lists the data slug to populate each cell.
          - We stop the loop at the first empty cell in that row.
          - ``DATA_SLUG_EMPTY_CELL`` allows to leave a column empty while continuing this loop.
        """
        # Loop over the sheets that must be populated to read their columns config.
        for sheet_index, sheet, config in self._sheets_iterator():
            # Loop over the sheet columns. Stop at the first empty cell.
            col = config["start_col"] - 1  # To start the loop with the increment, easier to read.
            row = config["start_row"]
            while True:
                col += 1

                # Get the data slug.
                data_slug = sheet.cell(column=col, row=row).value

                # Empty cell? => stop looking for config information.
                if not data_slug:
                    break

                # Remove the config information from the cell.
                sheet.cell(column=col, row=row).value = ""

                # Skip this column and continue to look for config information.
                if data_slug == DATA_SLUG_EMPTY_CELL:
                    continue

                # Register the data slug to use for this column.
                self._config[sheet_index]["columns"][col] = data_slug

                # Register this model/field for this sheet.
                self._update_models_fields(sheet_index, data_slug)

    def _init_config_sheets(self):
        """
        Read config for the sheets to populate.

        - The first row must contain column headings, and is therefore skipped.
        - The first column lists the index of the sheets that must be populated.
          - The first sheet without taking into account the config sheet has the index ``1``.
          - We stop the loop at the first empty cell in that column.
        - The second column lists the starting cell, using the cell name (e.g. ``B5``).
        """
        sheet = self.get_sheet(0)  # Get the config sheet.

        # Loop over rows starting from the second one (first row contains column headings).
        row = 1
        while True:
            row += 1

            # Get the sheet index from the first column. Stop at the first empty cell.
            sheet_index = sheet.cell(column=1, row=row).value

            # Empty cell? => stop looking for config information.
            if not sheet_index:
                break

            # Adapt the index as in the code counting starts at 0 (+ force int, not long).
            sheet_index = int(sheet_index) - 1

            # Convert the starting cell name into ``(col, row)`` indexes.
            start_col, start_row = self.cell_name_to_col_row(sheet.cell(column=2, row=row).value)

            # Store the config for this sheet.
            self._config[sheet_index] = {
                "start_col": start_col,
                "start_row": start_row,
                "columns": OrderedDict(),  # Populated in ``self._init_config_columns()``.
                "models_fields": OrderedDict(),  # Populated in ``self._init_config_columns()``.
            }

        # Remove the config sheet.
        self.book.remove_sheet(sheet)


class ExportExcelDetail(AbstractExportExcelTemplate):
    """Excel export for templates containing the data of a single record."""

    DEFAULT_FILENAME = "easynut-detail.xlsx"  # Default name for the file.
    DEFAULT_TEMPLATE = "easynut-detail.xlsx"  # Default template file to use.
    DEFAULT_LIST_CONFIG = {
        "col_inc": 1,
        "row_inc": 0,
        "max_values": 10,
    }

    def __init__(self, model, **kwargs):
        super(ExportExcelDetail, self).__init__(**kwargs)
        self.model = model

    def populate(self):
        """Populate the template with data."""
        super(ExportExcelDetail, self).populate()

        # Loop over sheets to populate.
        for sheet_index, sheet, config in self._sheets_iterator():
            # Loop over cells to populate.
            for cell_name, data_slug in config["cells"].iteritems():
                col, row = self.cell_name_to_col_row(cell_name)
                values = self.model.get_value_from_data_slug(data_slug)
                if type(values) not in (list, tuple):
                    sheet.cell(column=col, row=row).value = values
                else:
                    list_config = config["lists_config"].get(cell_name, self.DEFAULT_LIST_CONFIG)
                    max_values = list_config["max_values"]
                    list_col, list_row = col, row
                    for value in values[:max_values]:
                        sheet.cell(column=list_col, row=list_row).value = value
                        list_col += list_config["col_inc"]
                        list_row += list_config["row_inc"]

    def _init_config(self):
        """
        Read the template config.

        - See ``AbstractExportExcelTemplate`` for more specs.
        """
        self._config = OrderedDict()  # Reset the config.
        self._init_config_sheets()  # Read config for the sheets to populate.
        self._init_config_cells()  # Read config for the data to populate in each sheet.
        debug("CONFIG:", self._config)  # @DEBUG

    def _init_config_cells(self):
        """
        Read config for the data to populate in each sheet.

        - Data are populated in configured cells.
        - We scan every cell in the cell area for config information.
        """
        # Loop over the sheets that must be populated to scan for config information.
        for sheet_index, sheet, config in self._sheets_iterator():
            # Loop over the cell range.
            for col in range(config["start_col"], config["end_col"]):
                for row in range(config["start_row"], config["end_row"]):
                    # Get cell value (potential data slug).
                    data_slug = force_text(sheet.cell(column=col, row=row).value)

                    # If it's not a config information, skip it.
                    if RE_DATA_SLUG_VALIDATION.match(data_slug) is None:
                        continue

                    # Remove the config information from the cell.
                    sheet.cell(column=col, row=row).value = ""

                    # Register the data slug to use for this column.
                    cell_name = self.cell_col_row_to_name(col, row)
                    self._config[sheet_index]["cells"][cell_name] = data_slug

    def _init_config_lists(self, sheet):
        """
        Read config for lists of data.
        """
        START_COL = 4

        # Loop over rows starting from the second one (first row contains column headings).
        row = 1
        while True:
            row += 1

            # Get the sheet index from the first column. Stop at the first empty cell.
            sheet_index = sheet.cell(column=START_COL, row=row).value

            # Empty cell? => stop looking for config information.
            if not sheet_index:
                break

            # Adapt the index as in the code counting starts at 0 (+ force int, not long).
            sheet_index = int(sheet_index) - 1

            cell_name = sheet.cell(column=START_COL + 1, row=row).value
            self._config[sheet_index]["lists_config"][cell_name] = {
                "col_inc": int(sheet.cell(column=START_COL + 2, row=row).value),
                "row_inc": int(sheet.cell(column=START_COL + 3, row=row).value),
                "max_values": int(sheet.cell(column=START_COL + 4, row=row).value),
            }

    def _init_config_sheets(self):
        """
        Read config for the sheets to populate.

        - The first row must contain column headings, and is therefore skipped.
        - The first column lists the index of the sheets that must be populated.
          - The first sheet without taking into account the config sheet has the index ``1``.
          - We stop the loop at the first empty cell in that column.
        - The second column lists the cell range (e.g. ``A1:AN99``) to scan for config information.
        """
        sheet = self.get_sheet(0)  # Get the config sheet.

        # Loop over rows starting from the second one (first row contains column headings).
        row = 1
        while True:
            row += 1

            # Get the sheet index from the first column. Stop at the first empty cell.
            sheet_index = sheet.cell(column=1, row=row).value

            # Empty cell? => stop looking for config information.
            if not sheet_index:
                break

            # Adapt the index as in the code counting starts at 0 (+ force int, not long).
            sheet_index = int(sheet_index) - 1

            # Convert the cell range into ``(col, row)`` indexes.
            cell_range = sheet.cell(column=2, row=row).value.split(":")
            start_col, start_row = self.cell_name_to_col_row(cell_range[0])
            end_col, end_row = self.cell_name_to_col_row(cell_range[1])

            # Store the config for this sheet.
            self._config[sheet_index] = {
                "start_col": start_col,
                "start_row": start_row,
                "end_col": end_col,
                "end_row": end_row,
                "cells": OrderedDict(),  # Populated in ``self._init_config_cells()``.
                "lists_config": {},  # Populated in ``self._init_config_cells()``.
            }

        # Read config for lists of data.
        self._init_config_lists(sheet)

        # Remove the config sheet.
        self.book.remove_sheet(sheet)

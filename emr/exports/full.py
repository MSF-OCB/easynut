# -*- coding: utf-8 -*-
from openpyxl import Workbook

from ..models import NON_DYNAMIC_FIELDS_DB_COL_NAMES, DynamicRegistry

from .base import AbstractExportExcel


class ExportExcelFull(AbstractExportExcel):
    """Create an Excel file containing an export of the whole database."""

    DEFAULT_FILENAME = "easynut-full-export.xlsx"  # Default name for the file.
    VERBOSE = True  # Whether to include advanced information.

    def generate(self):
        """Generate the export."""
        self.book = Workbook()
        DynamicRegistry.load_models_config()
        verbose_style = self.get_style(self.STYLE_VERBOSE)

        # Loop over all models and fields config.
        for model_id, model_config in DynamicRegistry.models_config.iteritems():
            sheet = self.create_sheet(model_config.name)
            field_ids = []

            # Write headings.
            col = 0; row = 1  # NOQA
            for field_id, field_config in model_config.fields_config.iteritems():
                col += 1
                sheet.cell(column=col, row=row).value = field_config.name
                field_ids.append(field_id)

            if self.VERBOSE:
                col += 1  # Skip a column.
                for db_col_name in NON_DYNAMIC_FIELDS_DB_COL_NAMES:
                    col += 1
                    sheet.cell(column=col, row=row).value = db_col_name

            # Apply heading style and freeze heading row.
            style = self.get_style(self.STYLE_HEADING)
            max_col = max(col, self.APPLY_STYLE_MIN_NUM_COLS)
            self.apply_style(style, sheet, max_col, row, min_row=row)
            sheet.freeze_panes = sheet.cell(column=1, row=row + 1)

            # Write values.
            min_row = 2  # Skip heading row.
            row = min_row - 1    # To start the loops with the increment, easier to read.
            for model in model_config.objects.all():
                col = 0; row += 1  # NOQA
                for field_id in field_ids:
                    col += 1
                    value = self.protect_sensitive_data(model, field_id)
                    self.set_cell_value(sheet.cell(column=col, row=row), value)

                if self.VERBOSE:
                    col += 1  # Skip a column.
                    for db_col_name in NON_DYNAMIC_FIELDS_DB_COL_NAMES:
                        col += 1
                        self.set_cell_value(sheet.cell(column=col, row=row), getattr(model, db_col_name))

            # Apply verbose style.
            if self.VERBOSE:
                verbose_min_col = len(field_ids) + 1
                verbose_max_col = verbose_min_col + 1 + len(NON_DYNAMIC_FIELDS_DB_COL_NAMES)
                self.apply_style(verbose_style, sheet, verbose_max_col, row, min_col=verbose_min_col, min_row=min_row)

        # Remove the first sheet.
        self.book.remove_sheet(self.book.active)

        return self.book

    def _save_common(self):
        """Provide common steps for "save" methods."""
        if self.book is None:
            self.generate()
        super(ExportExcelFull, self)._save_common()
